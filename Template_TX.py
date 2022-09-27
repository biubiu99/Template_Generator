import numpy as np
from xlsxwriter import Workbook
from datetime import date
import pandas
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import string

# Create new Excel file
current_date = date.today()
current_date_modified = current_date.strftime("%Y-%m-%d")
Installation_date = current_date.strftime("%m/%d/%Y")

wb_TX_path = fr"C:\Users\Spareuser\Desktop\Code\template\Daily_templates\TX_14_{current_date_modified}_Valencia.xlsx"

wb_TX = Workbook(wb_TX_path)
wb_TX.close()

# Load SAP data
SAP_data_path = r"C:\Users\Spareuser\Desktop\Code\template\sample_data.xlsx"
data_loaded = pandas.read_excel(SAP_data_path)

# get 05 data for TODAY
TX_data = data_loaded[data_loaded["Warehouse"] == "14-US TX"]
TX_data_1 = TX_data[TX_data["Date of Installation"] == f"{Installation_date}"]
# get data for chairs only
TX_data_chairs = TX_data_1[~TX_data_1["Item/Service Description"].isin(
    ["Extended Warranty", "Ipad Holder", "White Glove Service", "Wine Holder", "Carbon Fiber Tray",
     "Accessory Package"])]

# get needed columns from original data
TX_data_modified = TX_data_chairs[
    ["Document Number", "Customer/Vendor Name", "Street", "City", "State", "Zip Code", "Phone Number", "Email",
     "Item/Service Description", "Quantity", "CC Notes", "Shipping Method"]]
# rename columns to match the template
TX_data_modified = TX_data_modified.rename(
    {"Customer/Vendor Name": "ShipToCompany", "Phone Number": "ShipToPhone", "Street": "ShipToAddress1",
     "Document Number": "ReferenceNumber",
     "Item/Service Description": "SKU", "CC Notes": "Carrier Notes", "Email": "ShipToEmail", "Zip Code": "ShipToZip",
     "State": "ShipToState", "City": "ShipToCity"}, axis=1)
# reset the index
TX_data_modified = TX_data_modified.reset_index(drop=False)

# adding columns to match the template
TX_data_modified = pandas.concat([TX_data_modified.iloc[:, :2],
                                  pandas.DataFrame('', columns=['PurchaseOrderNumber', 'ShipCarrier', 'ShipService',
                                                                'ShipBilling', "ShipAccount", "ShipDate", "CancelDate",
                                                                "Notes", "ShipTo Name"], index=TX_data_modified.index),
                                  TX_data_modified.iloc[:, 2:]], axis=1)

TX_data_modified.insert(loc=13, column="ShipToAddress2", value="")
TX_data_modified.insert(loc=17, column="ShipToCountry", value="USA")
TX_data_modified.insert(loc=19, column="ShipToFax", value="")
TX_data_modified = pandas.concat([TX_data_modified.iloc[:, :21],
                                  pandas.DataFrame("", columns=["ShipToCustomerID", "ShipToDeptNumber", "RetailerID"],
                                                   index=TX_data_modified.index), TX_data_modified.iloc[:, 21:]],
                                 axis=1)
TX_data_modified = pandas.concat([TX_data_modified.iloc[:, :26],
                                  pandas.DataFrame("", columns=["UseCOD", "UseInsurance", "Saved Elements",
                                                                "Order Item Saved Elements"],
                                                   index=TX_data_modified.index), TX_data_modified.iloc[:, 26:]],
                                 axis=1)

Shipping_method = TX_data_modified.loc[:, "Shipping Method"]
for i in range(Shipping_method.index[-1]):
    if TX_data_modified.loc[i, "Shipping Method"] == 1:
        TX_data_modified.at[i, "Carrier Notes"] = "white glove service needed"
    elif TX_data_modified.loc[i, "Shipping Method"] == 4:
        TX_data_modified.at[i, "Carrier Notes"] = "threshold service needed"
    elif TX_data_modified.loc[i, "Shipping Method"] == 3:
        continue

# drop the unecessary column
TX_data_modified.drop("index", inplace=True, axis=1)
TX_data_modified.drop("Shipping Method", inplace=True, axis=1)

with pandas.ExcelWriter(wb_TX_path, engine='xlsxwriter') as writer:
    # add empty row to the top of the worksheet
    TX_data_modified.to_excel(writer, index=False, startrow=1)
    # define worksheet and workbook
    workbook_14 = writer.book
    worksheet_14 = writer.sheets["Sheet1"]

    # change font style
    cell_format = workbook_14.add_format()
    cell_format.set_font_name("Cambria")
    worksheet_14.set_column("A:L", None, cell_format)
    # merge cells in first row
    merge_format = workbook_14.add_format({
        'bold': 1,
        'border': 0,
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': '#045DA8',
        "font_size": 28,
        "font_name": "Calibri",
        "font_color": "#FFFFFF"})
    merge_format_1 = workbook_14.add_format({
        'bold': 1,
        'border': 0,
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': '#045DA8',
        "font_size": 12,
        "font_name": "Calibri",
        "font_color": "#FFFFFF",
    })

    worksheet_14.merge_range("D1:F1", "Order Import Template", merge_format)
    worksheet_14.merge_range("A1:C1", "", merge_format)
    worksheet_14.merge_range("G1:AF1",
                             "*Please note that blue headers are the only fields required for import, but we highly recommend completing orange header values as well, as these fields are required for order",
                             merge_format_1)
    """
    format_green = workbook_14.add_format(
        {"bg_color": "#00B050", "font_size": 14, "font_color": "#FFFFFF", "bold": 1, "align": "center", 'border': 1})
    format_orange = workbook_14.add_format(
        {"bg_color": "#F79431", "font_size": 14, "font_color": "#FFFFFF", "bold": 1, "align": "center", 'border': 1})
    """

# use openpyxl to change the background color of the cell
wb_TX_1 = openpyxl.load_workbook(wb_TX_path)
wb_TX_sheet = wb_TX_1["Sheet1"]
font_style = Font(name="Calibri", size=14, color="FFFFFF", bold=1)
wb_TX_sheet["A2"].font = font_style
wb_TX_sheet['A2'].fill = PatternFill(patternType='solid', fgColor='00B050')

alphabet_string = string.ascii_lowercase
alphabet_list = list(alphabet_string)
print(alphabet_list)

for i in range(2, 5):
    wb_TX_sheet[f"{alphabet_list[i]}2"].font = font_style
    wb_TX_sheet[f"{alphabet_list[i]}2"].fill = PatternFill(patternType='solid', fgColor='F79431')

for i in range(10, 18):
    wb_TX_sheet[f"{alphabet_list[i]}2"].font = font_style
    wb_TX_sheet[f"{alphabet_list[i]}2"].fill = PatternFill(patternType='solid', fgColor='00B050')

wb_TX_sheet["T2"].font = font_style
wb_TX_sheet['T2'].fill = PatternFill(patternType='solid', fgColor='00B050')

wb_TX_sheet["X2"].font = font_style
wb_TX_sheet['X2'].fill = PatternFill(patternType='solid', fgColor='00B050')

wb_TX_sheet["Y2"].font = font_style
wb_TX_sheet['Y2'].fill = PatternFill(patternType='solid', fgColor='00B050')

alphabet_list = ['b', 'f', 'g', 'h', 'i', 'j', 's', 'u', 'v', 'w', 'z']

for element in alphabet_list:
    wb_TX_sheet[f"{element}2"].font = Font(name="Calibri", size=14, bold=1)
wb_TX_1.save(wb_TX_path)
